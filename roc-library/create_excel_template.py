#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
สร้าง Excel Template สำหรับแปลงข้อมูล VIP เป็น JSON
ใช้คำสั่ง: python create_excel_template.py
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import json
except ImportError:
    print("❌ ต้อง install openpyxl ก่อน: pip install openpyxl")
    exit(1)

def create_vip_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "VIP Data"
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Headers
    headers = ["Code", "Name (Thai)", "Facebook Name", "Facebook Link", "Phone"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add sample data
    sample_data = [
        ["VIP-1200", "วัชรพล จันทร์เรืองรอง", "Tak Kung", "https://www.facebook.com/sick.kz?mibextid=ZbWKwL", ""],
        ["VIP-1201", "รังสิรัตน์ เปรมฤดีสนิท", "Rangsirat Premrudesanit", "https://www.facebook.com/AuziieValentine", ""],
        ["VIP-1205", "ธิติสรรค์ แก้วซาว", "Tk Jame", "https://www.facebook.com/DevilJameZ", "0823176863"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Add empty rows for user input
    for row in range(len(sample_data) + 2, len(sample_data) + 20):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Create JSON output sheet
    json_ws = wb.create_sheet("JSON Output")
    json_ws.column_dimensions['A'].width = 100
    
    # Add instructions
    json_ws['A1'] = "JSON Output - Copy all text below and paste into your database"
    json_ws['A1'].font = Font(bold=True, size=12)
    json_ws['A2'] = "Steps:"
    json_ws['A3'] = "1. Fill in the VIP Data sheet with your information"
    json_ws['A4'] = "2. Run the Python script: python convert_vip_to_json.py"
    json_ws['A5'] = "3. Copy the generated JSON from vip_output.json"
    json_ws['A7'] = "[Example JSON output will appear here]"
    
    # Save
    wb.save('vip_converter_template.xlsx')
    print("✓ สร้างไฟล์ vip_converter_template.xlsx สำเร็จ")
    print("\n📋 วิธีใช้:")
    print("1. เปิดไฟล์ vip_converter_template.xlsx")
    print("2. กรอกข้อมูลใน Sheet 'VIP Data'")
    print("3. Copy ข้อมูล (รวมหัวแถว)")
    print("4. รัน Python script ด้วยคำสั่ง: python convert_vip_to_json.py")
    print("5. ตรวจสอบไฟล์ vip_output.json ที่จะสร้างขึ้น")

if __name__ == "__main__":
    create_vip_excel()
